#########################################################################
## Añade worksheets desde ficheros mensuales PyG_mes a                 ##
## fichero existente en PyG_anual si existe                            ##
## A principio de aó deberia generarse un fichero vacio con nombre     ##
## Nombre Comercio_Codigo Comercio_Año                                 ##
#########################################################################

import os
import pandas as pd
import locale
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Configuración
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RUTA_BASE = os.path.join(BASE_DIR, 'POC')
DIR_ENTRADA = os.path.join(RUTA_BASE, 'PyG_mes')
DIR_SALIDA = os.path.join(RUTA_BASE, 'PyG_anual')

def crear_directorio_si_no_existe(directorio):
    """
    Crea un directorio si no existe.
    """
    if not os.path.exists(directorio):
        os.makedirs(directorio)
        print(f"Directorio creado: {directorio}")
    else:
        print(f"Directorio existente: {directorio}")

def establecer_locale():
    """
    Establece el locale a español si es posible, de lo contrario usa el por defecto.
    """
    try:
        locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
    except locale.Error:
        try:
            locale.setlocale(locale.LC_ALL, 'Spanish_Spain.1252')  # Para Windows
        except locale.Error:
            print("No se pudo establecer el locale español. Usando el locale por defecto.")

def obtener_nombre_base(nombre_archivo: str) -> str:
    """
    Extraer el nombre base del archivo sin considerar el mes.
    """
    match = re.search(r"(.*)_\d{1,2}_PyG\.xlsx$", nombre_archivo)
    if match:
        return match.group(1)
    return None

def verificar_archivo_existe(ruta_archivo: str) -> bool:
    """
    Verifica si un archivo Excel existe y puede ser abierto.
    """
    try:
        if os.path.exists(ruta_archivo):
            # Intenta abrir el archivo para verificar que no está corrupto
            pd.read_excel(ruta_archivo, sheet_name=None)
            return True
        return False
    except Exception:
        return False

def copiar_worksheets(ruta_entrada: str, ruta_salida: str):
    """
    Copia los worksheets del archivo de entrada al archivo de salida, preservando fórmulas y estilos básicos
    y agregando nuevas hojas sin reemplazar las ya existentes en el archivo de salida.
    """
    try:
        # Cargar el libro de entrada y salida con openpyxl
        libro_entrada = load_workbook(ruta_entrada, data_only=False)  # `data_only=False` para mantener las fórmulas
        libro_salida = load_workbook(ruta_salida)

        # Iterar sobre cada hoja del libro de entrada
        for nombre_hoja in libro_entrada.sheetnames:
            hoja_entrada = libro_entrada[nombre_hoja]

            # Verificar si la hoja ya existe en el libro de salida
            if nombre_hoja in libro_salida.sheetnames:
                # Crear un nombre único si la hoja ya existe
                nuevo_nombre = f"{nombre_hoja}_copia"
                contador = 1
                while nuevo_nombre in libro_salida.sheetnames:
                    nuevo_nombre = f"{nombre_hoja}_copia_{contador}"
                    contador += 1
                hoja_salida = libro_salida.create_sheet(title=nuevo_nombre)
            else:
                hoja_salida = libro_salida.create_sheet(title=nombre_hoja)

            # Copiar todas las celdas (valores y fórmulas) de una hoja a otra
            for fila in hoja_entrada.iter_rows():
                for celda in fila:
                    nueva_celda = hoja_salida[celda.coordinate]
                    nueva_celda.value = celda.value  # Copiar valor y fórmula

            # Copiar anchos de columna y alturas de fila
            for col in hoja_entrada.column_dimensions:
                hoja_salida.column_dimensions[col] = hoja_entrada.column_dimensions[col]
            for row in hoja_entrada.row_dimensions:
                hoja_salida.row_dimensions[row] = hoja_entrada.row_dimensions[row]

            print(f"Añadida/actualizada hoja: {hoja_salida.title}")

        # Guardar los cambios en el archivo de salida
        libro_salida.save(ruta_salida)

    except Exception as e:
        print(f"Error al procesar el archivo {ruta_entrada}: {str(e)}")
        raise


def obtener_archivos_salida_existentes():
    """
    Obtiene un conjunto de nombres base de los archivos existentes en el directorio de salida.
    """
    archivos_salida = set()
    for archivo in os.listdir(DIR_SALIDA):
        if archivo.endswith('_PyG.xlsx'):
            nombre_base = archivo[:-9]  # Elimina '_PyG.xlsx'
            archivos_salida.add(nombre_base)
    return archivos_salida


def obtener_nombre_base(archivo):
    """
    Extrae el nombre base de un archivo (sin año, mes, ni extensión).
    Ejemplo: "Avda. de los Castros_4338_2024_9_PyG.xlsx" -> "Avda. de los Castros"
    """
    nombre_base = archivo.rsplit('_', 2)[0]  # Extraemos solo la parte antes del año y mes
    return nombre_base

def obtener_nombre_hoja(archivo_entrada):
    """
    Extrae el nombre de la hoja del archivo de entrada.
    Este ejemplo supone que el nombre de la hoja coincide con el mes/año del archivo de entrada.
    """
    # Asumimos que el nombre de la hoja sigue el formato "XX_PyG" donde XX es el mes o año
    nombre_hoja = archivo_entrada.split('_')[-2] + '_PyG'
    return nombre_hoja

def verificar_hoja_existente(ruta_archivo_salida, nombre_hoja):
    """
    Verifica si una hoja con el nombre dado ya existe en el archivo de salida.
    """
    try:
        wb = openpyxl.load_workbook(ruta_archivo_salida)
        return nombre_hoja in wb.sheetnames
    except Exception as e:
        print(f"Error al verificar hoja: {e}")
        return False

def main():
    """
    Función principal para ejecutar el proceso completo.
    """
    try:
        print("\nIniciando proceso de copia de hojas de cálculo...")
        establecer_locale()
        
        # Crear directorios si no existen
        print("\nVerificando directorios...")
        crear_directorio_si_no_existe(RUTA_BASE)
        crear_directorio_si_no_existe(DIR_ENTRADA)
        crear_directorio_si_no_existe(DIR_SALIDA)
        
        print(f"\nDirectorio de entrada: {DIR_ENTRADA}")
        print(f"Directorio de salida: {DIR_SALIDA}")

        # Obtener nombres base de archivos existentes en el directorio de salida
        nombres_base_salida = obtener_archivos_salida_existentes()
        print(f"\nArchivos existentes en salida: {nombres_base_salida}")

        archivos_procesados = 0
        # Procesar todos los archivos de entrada sin excluir archivos duplicados
        print("\nProcesando archivos...")
        for archivo_entrada in os.listdir(DIR_ENTRADA):
            if archivo_entrada.endswith('.xlsx'):
                nombre_base_entrada = obtener_nombre_base(archivo_entrada)
                print(f"\nProcesando archivo: {archivo_entrada}")

                # Verificar si el archivo de entrada tiene un archivo base correspondiente en la salida
                for nombre_base_salida in nombres_base_salida:
                    if nombre_base_entrada == nombre_base_salida:
                        ruta_archivo_entrada = os.path.join(DIR_ENTRADA, archivo_entrada)
                        ruta_archivo_salida = os.path.join(DIR_SALIDA, f"{nombre_base_salida}_PyG.xlsx")

                        if verificar_archivo_existe(ruta_archivo_entrada):
                            try:
                                # Extraer la hoja correspondiente al archivo de entrada
                                hoja_nombre = obtener_nombre_hoja(archivo_entrada)
                                print(f"Procesando hoja: {hoja_nombre}")

                                # Verificar si la hoja ya existe en el archivo de salida
                                if not verificar_hoja_existente(ruta_archivo_salida, hoja_nombre):
                                    # Copiar las hojas del archivo de entrada en el archivo de salida correspondiente
                                    copiar_worksheets(ruta_archivo_entrada, ruta_archivo_salida)
                                    print(f"✅ Procesado exitosamente: {archivo_entrada}")
                                    archivos_procesados += 1
                                else:
                                    print(f"⚠️ Hoja {hoja_nombre} ya existe en {ruta_archivo_salida}, omitiendo...")
                            except Exception as e:
                                print(f"❌ Error al procesar {archivo_entrada}: {str(e)}")
                        else:
                            print(f"❌ No se puede acceder al archivo: {archivo_entrada}")
                        break
                else:
                    print(f"⚠️ Ignorando {archivo_entrada}: No existe archivo correspondiente en el directorio de salida")

        print(f"\nProceso completado. Archivos procesados: {archivos_procesados}")

    except Exception as e:
        print(f"\n❌ Error en el proceso principal: {str(e)}")
        raise

if __name__ == "__main__":
    main()